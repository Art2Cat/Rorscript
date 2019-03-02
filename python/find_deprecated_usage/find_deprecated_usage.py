#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import re
from concurrent.futures.thread import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from find_deprecated_usage.parse_deprecated_lib import parse


def write_to_sheet(source: dict):
    """save result to excel
    :param source: result dict
    :type source: dict
    """
    book = Workbook()
    sheet = book.active
    i = 0
    for k, v in source.items():
        # key is used deprecated methods class name, v is deprecated class set
        sheet.cell(row=i + 1, column=1).value = k
        for c in v:
            sheet.cell(row=i + 1, column=2).value = c.lib_name
            sheet.cell(row=i + 1, column=3).value = c.name
            sheet.cell(row=i + 1, column=4).value = c.usage
            i += 1

    book.save("result.xlsx")


def find(file: Path, libs: set):
    # TODO: update output format
    potential_libs = set()
    rss = file.read_text(encoding="utf-8")
    for s in rss.split("\n"):
        for c in libs:
            r = "({})".format(c.full_name)
            if re.search(r, s) is not None:
                potential_libs.add(c)

    for s in rss.split("\n"):
        for c in potential_libs:
            usage = 0
            for m in c.methods:
                r = "({})".format(m)
                if re.search(r, s) is not None:
                    usage += 1
            c.usage = usage

    print("{} deprecated method usage: ".format(file.name))
    for v in potential_libs:
        print("used class: {} in lib: {} usage: {}".format(v.name, v.lib_name, v.usage))


def main(dir_path: Path):
    start_time = datetime.now()
    libs = parse()
    executor = ThreadPoolExecutor(max_workers=10)
    for path in dir_path.rglob("*.java"):
        executor.submit(find, path, libs)
        # find(path, libs)
    end_time = datetime.now()
    print("estimate time: {}".format(end_time - start_time))


if __name__ == "__main__":
    main(Path.cwd())
